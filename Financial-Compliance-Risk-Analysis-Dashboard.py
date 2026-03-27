from __future__ import annotations

import math
import os
import sys
from pathlib import Path

import matplotlib.pyplot as plt
from matplotlib.ticker import PercentFormatter
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


INPUT_FILE = "big4_financial_risk_compliance_master2.xlsx"
OUTPUT_FILE = "Total_Audit_vs_Fraud_Dashboard.xlsx"


# ---------- Theme ----------
NAVY = "0B1736"
TEAL = "23C4B5"
BLUE = "3568D8"
PURPLE = "7B3FE4"
ORANGE = "F59E0B"
SLATE = "475569"
CARD_BG = "F8FAFC"
SHEET_BG = "EEF2F7"
TEXT = "14213D"
MUTED = "64748B"
BORDER = "CBD5E1"
LIGHT_HEADER = "D5DCE6"
MINT = "B8EFE7"


# ---------- Utility styling helpers ----------
def fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def thin_border(color: str = BORDER) -> Border:
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def set_cell(ws, cell_ref, value=None, *, font=None, fill_color=None, align=None, border=None, fmt=None):
    c = ws[cell_ref]
    if value is not None:
        c.value = value
    if font is not None:
        c.font = font
    if fill_color is not None:
        c.fill = fill(fill_color)
    if align is not None:
        c.alignment = align
    if border is not None:
        c.border = border
    if fmt is not None:
        c.number_format = fmt
    return c


def merge_block(ws, start_row, end_row, start_col, end_col, value=None):
    ws.merge_cells(
        start_row=start_row,
        end_row=end_row,
        start_column=start_col,
        end_column=end_col,
    )
    c = ws.cell(start_row, start_col)
    if value is not None:
        c.value = value
    return c


def style_range(ws, cell_range: str, *, fill_color=None, font=None, align=None, border=None):
    for row in ws[cell_range]:
        for c in row:
            if fill_color is not None:
                c.fill = fill(fill_color)
            if font is not None:
                c.font = font
            if align is not None:
                c.alignment = align
            if border is not None:
                c.border = border


def draw_card(ws, col_start, col_end, title, formula, subtitle, title_color):
    merge_block(ws, 5, 5, col_start, col_end, title)
    merge_block(ws, 6, 7, col_start, col_end, formula)
    merge_block(ws, 8, 8, col_start, col_end, subtitle)

    title_cell = ws.cell(5, col_start)
    title_cell.font = Font(color="FFFFFF", bold=True, size=12)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    style_range(ws, f"{get_column_letter(col_start)}5:{get_column_letter(col_end)}5", fill_color=title_color)

    mid = f"{get_column_letter(col_start)}6:{get_column_letter(col_end)}7"
    style_range(
        ws,
        mid,
        fill_color=CARD_BG,
        border=thin_border(),
        align=Alignment(horizontal="center", vertical="center"),
    )
    main_cell = ws.cell(6, col_start)
    main_cell.font = Font(color=TEXT, bold=True, size=22)

    footer = f"{get_column_letter(col_start)}8:{get_column_letter(col_end)}8"
    style_range(
        ws,
        footer,
        fill_color=CARD_BG,
        border=thin_border(),
        align=Alignment(horizontal="center", vertical="center"),
    )
    foot_cell = ws.cell(8, col_start)
    foot_cell.font = Font(color=MUTED, italic=True, size=11)


def write_table(ws, start_row, start_col, data, col_widths=None, value_formats=None):
    for r_idx, row in enumerate(data, start=start_row):
        for c_idx, value in enumerate(row, start=start_col):
            cell = ws.cell(r_idx, c_idx, value)
            cell.border = thin_border("E2E8F0")
            if r_idx == start_row:
                cell.fill = fill(LIGHT_HEADER)
                cell.font = Font(bold=True, color=TEXT, size=12)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                if (r_idx - start_row) % 2 == 1:
                    cell.fill = fill("F8FAFC")
                else:
                    cell.fill = fill("EEF2F7")
                cell.font = Font(color=TEXT, size=11)
                if c_idx == start_col:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if value_formats and c_idx in value_formats:
                        cell.number_format = value_formats[c_idx]
    if col_widths:
        for offset, width in enumerate(col_widths):
            ws.column_dimensions[get_column_letter(start_col + offset)].width = width


def create_scatter_plot(x_vals, y_vals, intercept, slope, r2, out_path: Path):
    fig, ax = plt.subplots(figsize=(9.5, 5.0), dpi=160)
    ax.scatter(x_vals, y_vals, s=22, alpha=0.85, edgecolors="none")

    x_sorted = sorted(x_vals)
    y_fit = [intercept + slope * x for x in x_sorted]
    ax.plot(x_sorted, y_fit, linewidth=2)

    ax.set_xlabel("Total Audit Engagements", fontsize=10)
    ax.set_ylabel("Fraud Cases Detected", fontsize=10)
    ax.grid(True, linestyle="--", alpha=0.25)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)

    eq_text = f"y = {intercept:.2f} + {slope:.4f}x\nR square = {r2:.1%}"
    ax.text(
        0.02,
        0.97,
        eq_text,
        transform=ax.transAxes,
        va="top",
        ha="left",
        fontsize=9,
        bbox=dict(boxstyle="round,pad=0.25", facecolor="#E2E8F0", edgecolor="#CBD5E1"),
    )

    plt.tight_layout()
    fig.savefig(out_path, bbox_inches="tight", facecolor="white")
    plt.close(fig)


def create_histogram_plot(labels, freqs, cumulative, out_path: Path):
    fig, ax1 = plt.subplots(figsize=(9.5, 5.0), dpi=160)
    x = list(range(len(labels)))
    ax1.bar(x, freqs, width=0.72)
    ax1.set_xticks(x)
    ax1.set_xticklabels([str(v) for v in labels], fontsize=8)
    ax1.set_ylabel("Frequency", fontsize=10)
    ax1.set_xlabel("Fraud case bins", fontsize=10)
    ax1.grid(True, axis="y", linestyle="--", alpha=0.25)
    ax1.spines["top"].set_visible(False)
    ax1.spines["right"].set_visible(False)

    ax2 = ax1.twinx()
    ax2.plot(x, cumulative, marker="o", linewidth=2, markersize=3)
    ax2.yaxis.set_major_formatter(PercentFormatter(1.0))
    ax2.set_ylim(0, 1)
    ax2.spines["top"].set_visible(False)

    plt.tight_layout()
    fig.savefig(out_path, bbox_inches="tight", facecolor="white")
    plt.close(fig)


def build_dashboard(input_path: Path, output_path: Path):
    wb = load_workbook(input_path)

    for name in ["Regression Analysis", "Histogram", "Correlation & Covariance", "Descriptive Statistics"]:
        if name not in wb.sheetnames:
            raise ValueError(f"Required sheet missing: {name}")

    if "Dashboard" in wb.sheetnames:
        del wb["Dashboard"]
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False

    # Background and layout
    for col in range(1, 19):
        ws.column_dimensions[get_column_letter(col)].width = 15
    ws.row_dimensions[1].height = 34
    ws.row_dimensions[3].height = 26
    ws.row_dimensions[5].height = 24
    ws.row_dimensions[6].height = 36
    ws.row_dimensions[7].height = 34
    ws.row_dimensions[8].height = 20
    for r in [10, 24, 36]:
        ws.row_dimensions[r].height = 24

    style_range(ws, "A1:R41", fill_color=SHEET_BG)

    # Header
    merge_block(ws, 1, 2, 1, 18, "Total Audit vs Fraud Dashboard")
    style_range(ws, "A1:R2", fill_color=NAVY)
    set_cell(
        ws,
        "A1",
        font=Font(color="FFFFFF", bold=True, size=28),
        align=Alignment(horizontal="left", vertical="center"),
    )
    merge_block(
        ws,
        3,
        3,
        1,
        18,
        "Big 4 Financial Risk Compliance | Built only from Regression Analysis, Histogram, Correlation & Covariance, and Descriptive Statistics",
    )
    style_range(ws, "A3:R3", fill_color=NAVY)
    set_cell(
        ws,
        "A3",
        font=Font(color="D1D5DB", size=13),
        align=Alignment(horizontal="left", vertical="center"),
    )
    style_range(ws, "A4:R4", fill_color=TEAL)
    ws.row_dimensions[4].height = 8

    # KPI cards
    cards = [
        (1, 3, "Observations", "='Regression Analysis'!B8", "Regression sample size", BLUE),
        (4, 6, "Correlation (r)", "='Correlation & Covariance'!E4", "Weak positive link", TEAL),
        (7, 9, "R Square", "='Regression Analysis'!B5", "Explained variance", PURPLE),
        (10, 12, "Significance F", "='Regression Analysis'!F12", "Overall model p-value", ORANGE),
        (13, 15, "Avg Audit Engagements", "='Descriptive Statistics'!E4", "Descriptive mean", SLATE),
        (16, 18, "Avg Fraud Cases", "='Descriptive Statistics'!G4", "Descriptive mean", BLUE),
    ]
    for c in cards:
        draw_card(ws, *c)

    ws["A6"].number_format = "0"
    ws["D6"].number_format = "0.000"
    ws["G6"].number_format = "0.0%"
    ws["J6"].number_format = "0.0000"
    ws["M6"].number_format = "#,##0.0"
    ws["P6"].number_format = "0.0"

    # Section headers
    merge_block(ws, 10, 10, 1, 9, "Relationship View | Correlation + Regression")
    merge_block(ws, 10, 10, 10, 18, "Distribution View | Histogram")
    style_range(ws, "A10:I10", fill_color=NAVY)
    style_range(ws, "J10:R10", fill_color=NAVY)
    set_cell(ws, "A10", font=Font(color="FFFFFF", bold=True, size=14), align=Alignment(horizontal="left"))
    set_cell(ws, "J10", font=Font(color="FFFFFF", bold=True, size=14), align=Alignment(horizontal="left"))

    merge_block(ws, 24, 24, 1, 6, "Descriptive Statistics Snapshot")
    merge_block(ws, 24, 24, 7, 12, "Regression Summary")
    merge_block(ws, 24, 24, 13, 18, "Correlation & Covariance Snapshot")
    style_range(ws, "A24:F24", fill_color=NAVY)
    style_range(ws, "G24:L24", fill_color=NAVY)
    style_range(ws, "M24:R24", fill_color=NAVY)
    for ref in ["A24", "G24", "M24"]:
        set_cell(ws, ref, font=Font(color="FFFFFF", bold=True, size=14), align=Alignment(horizontal="left"))

    # Chart image generation from source-sheet values
    reg = wb["Regression Analysis"]
    hist = wb["Histogram"]
    corr = wb["Correlation & Covariance"]
    desc = wb["Descriptive Statistics"]

    x_vals = [corr[f"A{r}"].value for r in range(2, 102) if corr[f"A{r}"].value is not None]
    y_vals = [corr[f"B{r}"].value for r in range(2, 102) if corr[f"B{r}"].value is not None]
    intercept = reg["B17"].value
    slope = reg["B18"].value
    r2 = reg["B5"].value

    labels = [hist[f"A{r}"].value for r in range(2, 13)]
    freqs = [hist[f"B{r}"].value for r in range(2, 13)]
    cumulative = [hist[f"C{r}"].value for r in range(2, 13)]

    temp_dir = output_path.parent / "dashboard_assets"
    temp_dir.mkdir(parents=True, exist_ok=True)
    scatter_png = temp_dir / "scatter.png"
    hist_png = temp_dir / "histogram.png"

    create_scatter_plot(x_vals, y_vals, intercept, slope, r2, scatter_png)
    create_histogram_plot(labels, freqs, cumulative, hist_png)

    scatter_img = Image(str(scatter_png))
    scatter_img.width = 730
    scatter_img.height = 340
    scatter_img.anchor = "A11"
    ws.add_image(scatter_img)

    hist_img = Image(str(hist_png))
    hist_img.width = 730
    hist_img.height = 340
    hist_img.anchor = "J11"
    ws.add_image(hist_img)

    # Table blocks
    desc_table = [
        ["Metric", "Audit", "Fraud"],
        ["Mean", "='Descriptive Statistics'!E4", "='Descriptive Statistics'!G4"],
        ["Median", "='Descriptive Statistics'!E6", "='Descriptive Statistics'!G6"],
        ["Std Deviation", "='Descriptive Statistics'!E8", "='Descriptive Statistics'!G8"],
        ["Minimum", "='Descriptive Statistics'!E13", "='Descriptive Statistics'!G13"],
        ["Maximum", "='Descriptive Statistics'!E14", "='Descriptive Statistics'!G14"],
        ["Range", "='Descriptive Statistics'!E12", "='Descriptive Statistics'!G12"],
        ["Count", "='Descriptive Statistics'!E16", "='Descriptive Statistics'!G16"],
    ]
    write_table(ws, 25, 1, desc_table, col_widths=[18, 12, 12], value_formats={2: "#,##0.0", 3: "0.0"})
    for r in range(26, 33):
        ws[f"B{r}"].number_format = "#,##0.0"
        ws[f"C{r}"].number_format = "0.0"
        if r in [27, 29, 30, 31, 32]:
            ws[f"B{r}"].number_format = "#,##0"
            ws[f"C{r}"].number_format = "0"

    reg_table = [
        ["Metric", "Value"],
        ["Multiple R", "='Regression Analysis'!B4"],
        ["R Square", "='Regression Analysis'!B5"],
        ["Adjusted R Square", "='Regression Analysis'!B6"],
        ["Std Error", "='Regression Analysis'!B7"],
        ["F Statistic", "='Regression Analysis'!E12"],
        ["Significance F", "='Regression Analysis'!F12"],
        ["Intercept", "='Regression Analysis'!B17"],
        ["Audit Coefficient", "='Regression Analysis'!B18"],
        ["Coeff. P-value", "='Regression Analysis'!E18"],
        ["Prediction Formula", '="Predicted Fraud = " & TEXT(B31,"0.00") & " + " & TEXT(B32,"0.0000") & " x Total Audit Engagements"'],
    ]
    write_table(ws, 25, 7, reg_table, col_widths=[20, 28], value_formats={8: "0.000", 9: "0.0%"})
    for r in [26, 30]:
        ws[f"H{r}"].number_format = "0.000"
    for r in [27, 28]:
        ws[f"H{r}"].number_format = "0.0%"
    ws["H29"].number_format = "0.00"
    ws["H31"].number_format = "0.00"
    ws["H32"].number_format = "0.0000"
    ws["H33"].number_format = "0.0000"
    ws["H35"].alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")

    # Correlation matrix snapshot
    merge_block(ws, 25, 25, 13, 18, "Correlation matrix")
    style_range(ws, "M25:R25", fill_color="F8FAFC")
    set_cell(ws, "M25", font=Font(bold=True, color="334155", size=12), align=Alignment(horizontal="left"))

    # Header pills
    merge_block(ws, 26, 26, 15, 16, "Audit")
    merge_block(ws, 26, 26, 17, 18, "Fraud")
    merge_block(ws, 27, 27, 13, 13, "Audit")
    merge_block(ws, 28, 28, 13, 13, "Fraud")
    style_range(ws, "O26:P26", fill_color=LIGHT_HEADER, border=thin_border(), align=Alignment(horizontal="center"))
    style_range(ws, "Q26:R26", fill_color=LIGHT_HEADER, border=thin_border(), align=Alignment(horizontal="center"))
    style_range(ws, "M27:M27", fill_color=LIGHT_HEADER, border=thin_border(), align=Alignment(horizontal="center"))
    style_range(ws, "M28:M28", fill_color=LIGHT_HEADER, border=thin_border(), align=Alignment(horizontal="center"))
    for ref in ["O26", "Q26", "M27", "M28"]:
        set_cell(ws, ref, font=Font(bold=True, size=12, color=TEXT), align=Alignment(horizontal="center", vertical="center"))

    merge_block(ws, 27, 27, 15, 16, "=1")
    merge_block(ws, 27, 27, 17, 18, "='Correlation & Covariance'!E4")
    merge_block(ws, 28, 28, 15, 16, "='Correlation & Covariance'!E4")
    merge_block(ws, 28, 28, 17, 18, "=1")
    style_range(ws, "O27:P27", fill_color=TEAL, border=thin_border(), align=Alignment(horizontal="center"))
    style_range(ws, "Q27:R27", fill_color=MINT, border=thin_border(), align=Alignment(horizontal="center"))
    style_range(ws, "O28:P28", fill_color=MINT, border=thin_border(), align=Alignment(horizontal="center"))
    style_range(ws, "Q28:R28", fill_color=TEAL, border=thin_border(), align=Alignment(horizontal="center"))
    for ref in ["O27", "Q27", "O28", "Q28"]:
        set_cell(ws, ref, font=Font(bold=True, size=13, color=("FFFFFF" if ref in ["O27", "Q28"] else TEXT)))
        ws[ref].number_format = "0.000"

    merge_block(ws, 30, 30, 13, 15, "Covariance")
    merge_block(ws, 30, 30, 16, 18, "='Correlation & Covariance'!E9")
    merge_block(ws, 31, 31, 13, 15, "Var(Audit)")
    merge_block(ws, 31, 31, 16, 18, "='Correlation & Covariance'!E8")
    merge_block(ws, 32, 32, 13, 15, "Var(Fraud)")
    merge_block(ws, 32, 32, 16, 18, "='Correlation & Covariance'!F9")
    for rng in ["M30:O30", "M31:O31", "M32:O32"]:
        style_range(ws, rng, fill_color="F8FAFC", border=thin_border(), align=Alignment(horizontal="left"))
    for rng in ["P30:R30", "P31:R31", "P32:R32"]:
        style_range(ws, rng, fill_color="F8FAFC", border=thin_border(), align=Alignment(horizontal="right"))
    for ref in ["M30", "M31", "M32"]:
        set_cell(ws, ref, font=Font(color="334155", size=12))
    for ref in ["P30", "P31", "P32"]:
        set_cell(ws, ref, font=Font(color=TEXT, bold=True, size=12))
        ws[ref].number_format = "#,##0.0"

    merge_block(
        ws,
        33,
        34,
        13,
        18,
        'Positive covariance and a 0.267 correlation suggest fraud detections generally rise as audit engagement volume increases.',
    )
    style_range(ws, "M33:R34", fill_color="F8FAFC", border=thin_border(), align=Alignment(horizontal="left", vertical="center", wrap_text=True))
    set_cell(ws, "M33", font=Font(size=11, color=MUTED))

    # Insight cards
    insights = [
        (
            1,
            6,
            "Insight 1",
            '="The relationship is positive but modest: r = " & TEXT(D6,"0.000") & ", indicating a weak upward association between audit volume and fraud detection."',
            BLUE,
            "DCE8FF",
        ),
        (
            7,
            12,
            "Insight 2",
            '="The regression is statistically significant (p = " & TEXT(J6,"0.0000") & "), but R square is only " & TEXT(G6,"0.0%") & ", so audit volume alone explains a limited share of variation."',
            PURPLE,
            "ECE3FF",
        ),
        (
            13,
            18,
            "Insight 3",
            '="Fraud detections are widely dispersed (min " & TEXT(C29,"0") & ", max " & TEXT(C30,"0") & ", standard deviation " & TEXT(C28,"0.0") & "), showing materially different risk profiles across observations."',
            TEAL,
            "D9F7F2",
        ),
    ]
    for start_col, end_col, title, body, head_color, body_color in insights:
        merge_block(ws, 36, 36, start_col, end_col, title)
        style_range(ws, f"{get_column_letter(start_col)}36:{get_column_letter(end_col)}36", fill_color=head_color)
        set_cell(ws, f"{get_column_letter(start_col)}36", font=Font(color="FFFFFF", bold=True, size=14), align=Alignment(horizontal="left"))
        merge_block(ws, 37, 39, start_col, end_col, body)
        style_range(
            ws,
            f"{get_column_letter(start_col)}37:{get_column_letter(end_col)}39",
            fill_color=body_color,
            border=thin_border(),
            align=Alignment(horizontal="left", vertical="top", wrap_text=True),
        )
        set_cell(ws, f"{get_column_letter(start_col)}37", font=Font(color="475569", size=11))

    # Freeze panes and save
    ws.freeze_panes = "A5"
    wb.save(output_path)


if __name__ == "__main__":
    base_dir = Path.cwd()
    input_path = Path(sys.argv[1]) if len(sys.argv) > 1 else base_dir / INPUT_FILE
    output_path = Path(sys.argv[2]) if len(sys.argv) > 2 else base_dir / OUTPUT_FILE
    build_dashboard(input_path, output_path)
    print(f"Dashboard workbook created: {output_path}")
